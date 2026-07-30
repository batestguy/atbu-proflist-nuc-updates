"""
import_export.py — Excel import/export logic
ATBU Academic Planning Portal — Phase 1 Foundation

Handles:
  1. Import from NUC format (10 columns: S/No, Last Name, First Name, ...)
  2. Import from Newly Promoted format (9 columns: S/N, NAME, SEX, ...)
  3. Export to NUC format (same 10 columns, with regenerated S/No.)
  4. Phone number parsing (split on '/', store multiple)
  5. Duplicate detection via composite UNIQUE key (UPDATE on conflict)
  6. Faculty name normalization
  7. Date normalization (various formats → YYYY-MM-DD)
"""

import os
import json
import re
from datetime import datetime, timezone
from typing import Dict, List, Optional, Tuple

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy.exc import IntegrityError

from database import (
    Professor, PhoneNumber, ImportHistory, AppSetting,
    get_session
)
from name_parser import split_full_name, parse_nuc_date, get_year_from_date


# ============================================================================
# FACULTY NORMALIZATION
# ============================================================================

def load_faculty_map(session) -> Dict[str, str]:
    """
    Load the faculty name normalization mapping from app_settings.
    
    Discovered from actual data scan:
      - "Environmental Technoloty" → "Environmental Technology"  (typo)
      - "College of Medical Science" → "College of Medical Sciences" (plural)
      - "Management Science" → "Management Sciences" (inconsistent)
      - "Engineering and Engineering technology" → "... Technology" (capitalization)
    """
    setting = session.query(AppSetting).filter_by(key="faculty_normalization").first()
    if setting:
        return json.loads(setting.value)
    return {}


def normalize_faculty(name: str, faculty_map: Dict[str, str]) -> Tuple[str, bool]:
    """
    Normalize a faculty name using the mapping.
    
    Args:
        name: Raw faculty name from Excel
        faculty_map: Dict of {incorrect: correct} mappings
    
    Returns:
        Tuple of (normalized_name, was_modified)
    """
    name = name.strip()
    if name in faculty_map:
        return faculty_map[name], True
    return name, False


# ============================================================================
# PHONE PARSING
# ============================================================================

def parse_phones(phone_str: Optional[str]) -> List[str]:
    """
    Parse phone numbers — split on common delimiters.
    
    From data analysis: 2 out of 172 records have multiple phones
    delimited by '/' (e.g., "0903378417/08026796700").
    
    Args:
        phone_str: Raw phone string from Excel (may contain multiple numbers)
    
    Returns:
        List of individual phone numbers
    """
    if not phone_str or not phone_str.strip():
        return []
    
    phone_str = phone_str.strip()
    
    # Split by common delimiters: / , ; or whitespace-separated numbers
    # First try splitting by / (most common in the data)
    parts = re.split(r'[/,;]', phone_str)
    
    phones = []
    for part in parts:
        part = part.strip()
        if part and len(part) >= 7:  # Minimum reasonable phone length
            phones.append(part)
    
    return phones if phones else [phone_str]


# ============================================================================
# NUC FORMAT IMPORT
# ============================================================================

def import_nuc_format(filepath: str) -> Dict[str, int]:
    """
    Import professors from the NUC format Excel file.
    
    Expected columns (from row 2):
      S/No, Last Name, First Name, Other Names, Date of Professorship,
      Faculty, Department, Area of Specialization, E-mail, Phone Number
    
    Args:
        filepath: Path to the .xlsx file
    
    Returns:
        Dict with keys: added, updated, skipped, total
    """
    session = get_session()
    faculty_map = load_faculty_map(session)
    result = {"added": 0, "updated": 0, "skipped": 0, "total": 0}
    
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
        ws = wb.active
        
        row_count = 0
        for row in ws.iter_rows(min_row=3, values_only=True):
            # Skip empty rows
            if not row or all(v is None for v in row):
                continue
            
            # Unpack columns (10 columns, 0-indexed)
            s_no = row[0]          # S/No (not stored)
            last_name = str(row[1]).strip() if row[1] else ""
            first_name = str(row[2]).strip() if row[2] else ""
            other_names = str(row[3]).strip() if row[3] else ""
            date_raw = str(row[4]).strip() if row[4] else ""
            faculty_raw = str(row[5]).strip() if row[5] else ""
            department = str(row[6]).strip() if row[6] else ""
            specialization = str(row[7]).strip() if row[7] else ""
            email = str(row[8]).strip() if row[8] else ""
            phone_raw = str(row[9]).strip() if row[9] else ""
            
            # Validate required fields
            if not last_name or not first_name or not department:
                result["skipped"] += 1
                continue
            
            # Handle datetime objects from openpyxl (Excel date-formatted cells)
            if isinstance(date_raw, datetime):
                date_str = date_raw.strftime("%Y-%m-%d")
            else:
                date_str = str(date_raw).strip() if date_raw else ""
            
            # Normalize
            date_normalized = parse_nuc_date(date_str)
            faculty_normalized, _ = normalize_faculty(faculty_raw, faculty_map)
            phones = parse_phones(phone_raw)
            added_year = get_year_from_date(date_normalized)
            
            # Check for duplicate via composite unique key
            existing = session.query(Professor).filter_by(
                last_name=last_name,
                first_name=first_name,
                department=department
            ).first()
            
            if existing:
                # UPDATE existing record (merge newer data)
                existing.other_names = other_names or existing.other_names
                existing.date_of_professorship = date_normalized or existing.date_of_professorship
                existing.faculty = faculty_normalized or existing.faculty
                existing.area_of_specialization = specialization or existing.area_of_specialization
                existing.email = email or existing.email
                existing.added_year = added_year or existing.added_year
                existing.source_file = os.path.basename(filepath)
                existing.updated_at = datetime.now(timezone.utc)
                session.flush()
                
                # Update phones: only add new ones, avoid duplicates
                existing_phones = set(p.phone for p in existing.phone_numbers)
                for phone in phones:
                    if phone not in existing_phones:
                        session.add(PhoneNumber(
                            professor_id=existing.id,
                            phone=phone,
                            is_primary=1 if not existing_phones else 0
                        ))
                result["updated"] += 1
            else:
                # INSERT new record
                prof = Professor(
                    last_name=last_name,
                    first_name=first_name,
                    other_names=other_names,
                    date_of_professorship=date_normalized,
                    faculty=faculty_normalized,
                    department=department,
                    area_of_specialization=specialization,
                    email=email,
                    added_year=added_year,
                    source_file=os.path.basename(filepath),
                )
                session.add(prof)
                session.flush()  # Get the ID
                
                # Add phone numbers
                for i, phone in enumerate(phones):
                    session.add(PhoneNumber(
                        professor_id=prof.id,
                        phone=phone,
                        is_primary=1 if i == 0 else 0
                    ))
                result["added"] += 1
            
            row_count += 1
        
        result["total"] = row_count
        
        # Log import in history
        session.add(ImportHistory(
            filename=os.path.basename(filepath),
            records_added=result["added"],
            records_updated=result["updated"],
            records_skipped=result["skipped"],
        ))
        session.commit()
        
    except Exception as e:
        session.rollback()
        raise RuntimeError(f"Import failed for {filepath}: {e}")
    finally:
        session.close()
    
    return result


# ============================================================================
# NEWLY PROMOTED FORMAT IMPORT
# ============================================================================

def import_newly_promoted(filepath: str) -> Dict[str, int]:
    """
    Import professors from the Newly Promoted format Excel file.
    
The file has 3 sections with different column layouts:
      1. Newly Promoted (main section):
         S/N, NAME, SEX, D.O.L.P, RANK, AREA OF SPECIALIZATION,
         PHONE NUMBERS, EMAIL-ADDRESS, DEPARTMENT
      2. Retired section (different columns!):
         S/N, Last Name, First Name, Other Names, RANK, DATE, STATUS, DEPARTMENT
      3. Associate Professors (same as main section):
         S/N, NAME, SEX, D.O.L.P, RANK, AREA OF SPECIALIZATION,
         PHONE NUMBERS, EMAIL-ADDRESS, DEPARTMENT
    
    Args:
        filepath: Path to the .xlsx file
    
    Returns:
        Dict with keys: added, updated, skipped, total
    """
    session = get_session()
    faculty_map = load_faculty_map(session)
    result = {"added": 0, "updated": 0, "skipped": 0, "total": 0}
    
    # Status keywords that indicate this is a retired professor row
    RETIRED_STATUSES = {"retirement", "death", "transfer of service"}
    
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
        ws = wb.active
        
        # Detect sections by scanning all rows
        # We'll track which section we're in
        current_section = "main"  # "main", "retired", "associate"
        row_count = 0
        
        for row_idx in range(1, ws.max_row + 1):
            row_vals = []
            for col_idx in range(1, min(12, ws.max_column + 1)):
                row_vals.append(ws.cell(row=row_idx, column=col_idx).value)
            
            # Skip completely empty rows
            if all(v is None for v in row_vals):
                continue
            
            # Get first few cells as strings
            c1 = str(row_vals[0] or "").strip()
            c2 = str(row_vals[1] or "").strip() if len(row_vals) > 1 else ""
            c3 = str(row_vals[2] or "").strip() if len(row_vals) > 2 else ""
            c4 = str(row_vals[3] or "").strip() if len(row_vals) > 3 else ""
            
            # Detect section headers
            c1_lower = c1.lower()
            c2_lower = c2.lower()
            
            if "list of" in c1_lower or "list of" in c2_lower:
                # Section header — detect which section
                combined = (c1 + " " + c2 + " " + c3 + " " + c4).lower()
                if "retire" in combined:
                    current_section = "retired"
                elif "associate" in combined:
                    current_section = "associate"
                else:
                    current_section = "main"
                continue
            
            # Skip title rows (university name, office name, etc.)
            if "abubakar" in c1_lower or "office" in c1_lower or "university" in c1_lower:
                continue
            
            # Skip column header rows
            if c1 in ("S/N", "S/N0", "SN", "S.No") and ("name" in c2_lower or "s/n" in c2_lower):
                continue
            
            # Skip if this looks like a column header (NAME, SEX, etc.)
            if c2_lower == "name" and c3_lower in ("sex", "m", "f"):
                continue
            
            # Skip if S/N is not a number (header row indicator)
            try:
                s_no = int(c1)
            except (ValueError, TypeError):
                # Not a data row — could be a header or section title
                if any(keyword in c1_lower for keyword in ["list", "professor", "effective"]):
                    continue
                # Check if it's a name in the retired section (Col2=First, Col3=Other, Col4=Last)
                if current_section == "retired" and c2:
                    pass  # Process as retired row below
                else:
                    continue
            
            # ══════════════════════════════════════════════════════════════════
            # RETIRED SECTION: Different column layout!
            # Col1=S/N, Col2=Last Name, Col3=First Name, Col4=Other Names,
            # Col5=RANK, Col6=DATE, Col7=STATUS, Col8=DEPARTMENT
            # ══════════════════════════════════════════════════════════════════
            if current_section == "retired":
                last_name = str(row_vals[1] or "").strip() if len(row_vals) > 1 else ""
                first_name = str(row_vals[2] or "").strip() if len(row_vals) > 2 else ""
                other_names = str(row_vals[3] or "").strip() if len(row_vals) > 3 else ""
                rank = str(row_vals[4] or "Professor").strip() if len(row_vals) > 4 else "Professor"
                date_raw = str(row_vals[5] or "").strip() if len(row_vals) > 5 else ""
                retirement_status = str(row_vals[6] or "").strip() if len(row_vals) > 6 else ""
                department = str(row_vals[7] or "").strip() if len(row_vals) > 7 else ""
                
                # Validate
                if not last_name or not first_name:
                    result["skipped"] += 1
                    continue
                
                # Handle datetime objects
                if isinstance(date_raw, datetime):
                    date_str = date_raw.strftime("%Y-%m-%d")
                else:
                    date_str = date_raw
                
                date_normalized = parse_nuc_date(date_str)
                added_year = get_year_from_date(date_normalized)
                
                # Determine if this is a retirement, death, or transfer
                is_retired = 1 if retirement_status.lower() in RETIRED_STATUSES else 0
                
                # Faculty: not present in retired section
                faculty = ""
                
                # Check for duplicate
                existing = session.query(Professor).filter_by(
                    last_name=last_name,
                    first_name=first_name,
                    department=department
                ).first()
                
                if existing:
                    existing.other_names = other_names or existing.other_names
                    existing.date_of_professorship = date_normalized or existing.date_of_professorship
                    existing.is_retired = is_retired
                    existing.retirement_status = retirement_status or existing.retirement_status
                    existing.retirement_date = date_normalized if is_retired else existing.retirement_date
                    existing.rank = rank or existing.rank
                    existing.added_year = added_year or existing.added_year
                    existing.source_file = os.path.basename(filepath)
                    existing.updated_at = datetime.now(timezone.utc)
                    session.flush()
                    result["updated"] += 1
                else:
                    prof = Professor(
                        last_name=last_name,
                        first_name=first_name,
                        other_names=other_names,
                        date_of_professorship=date_normalized,
                        faculty=faculty,
                        department=department,
                        area_of_specialization="",
                        rank=rank,
                        is_retired=is_retired,
                        retirement_status=retirement_status if is_retired else None,
                        retirement_date=date_normalized if is_retired else None,
                        added_year=added_year,
                        source_file=os.path.basename(filepath),
                    )
                    session.add(prof)
                    session.flush()
                    result["added"] += 1
                
                row_count += 1
                continue
            
            # ══════════════════════════════════════════════════════════════════
            # MAIN / ASSOCIATE SECTION: Standard layout
            # Col1=S/N, Col2=NAME, Col3=SEX, Col4=D.O.L.P, Col5=RANK,
            # Col6=AREA OF SPECIALIZATION, Col7=PHONE, Col8=EMAIL, Col9=DEPARTMENT
            # ══════════════════════════════════════════════════════════════════
            name_raw = c2
            sex = str(row_vals[2] or "").strip() if len(row_vals) > 2 else ""
            date_raw = str(row_vals[3] or "").strip() if len(row_vals) > 3 else ""
            rank = str(row_vals[4] or "Professor").strip() if len(row_vals) > 4 else "Professor"
            specialization = str(row_vals[5] or "").strip() if len(row_vals) > 5 else ""
            phone_raw = str(row_vals[6] or "").strip() if len(row_vals) > 6 else ""
            email = str(row_vals[7] or "").strip() if len(row_vals) > 7 else ""
            department = str(row_vals[8] or "").strip() if len(row_vals) > 8 else ""
            
            # Validate
            if not name_raw:
                result["skipped"] += 1
                continue
            
            # Split name
            name_parts = split_full_name(name_raw)
            if not name_parts["last_name"] and not name_parts["first_name"]:
                result["skipped"] += 1
                continue
            
            # Handle datetime objects from openpyxl
            if isinstance(date_raw, datetime):
                date_str = date_raw.strftime("%Y-%m-%d")
            else:
                date_str = str(date_raw).strip() if date_raw else ""
            
            # Parse date
            date_normalized = parse_nuc_date(date_str)
            added_year = get_year_from_date(date_normalized)
            
            # Phone parsing
            phones = parse_phones(phone_raw)
            
            # Faculty: not present in this format, leave blank or infer later
            faculty = ""
            
            # Check for duplicate via composite unique key
            existing = session.query(Professor).filter_by(
                last_name=name_parts["last_name"],
                first_name=name_parts["first_name"],
                department=department
            ).first()
            
            if existing:
                # UPDATE
                existing.other_names = name_parts["other_names"] or existing.other_names
                existing.date_of_professorship = date_normalized or existing.date_of_professorship
                existing.area_of_specialization = specialization or existing.area_of_specialization
                existing.email = email or existing.email
                existing.sex = sex or existing.sex
                existing.rank = rank or existing.rank
                existing.added_year = added_year or existing.added_year
                existing.source_file = os.path.basename(filepath)
                existing.updated_at = datetime.now(timezone.utc)
                session.flush()
                
                # Merge phones
                existing_phones = set(p.phone for p in existing.phone_numbers)
                for phone in phones:
                    if phone not in existing_phones:
                        session.add(PhoneNumber(
                            professor_id=existing.id,
                            phone=phone,
                            is_primary=1 if not existing_phones else 0
                        ))
                result["updated"] += 1
            else:
                # INSERT
                prof = Professor(
                    last_name=name_parts["last_name"],
                    first_name=name_parts["first_name"],
                    other_names=name_parts["other_names"],
                    date_of_professorship=date_normalized,
                    faculty=faculty,
                    department=department,
                    area_of_specialization=specialization,
                    email=email,
                    sex=sex,
                    rank=rank,
                    added_year=added_year,
                    source_file=os.path.basename(filepath),
                )
                session.add(prof)
                session.flush()
                
                for i, phone in enumerate(phones):
                    session.add(PhoneNumber(
                        professor_id=prof.id,
                        phone=phone,
                        is_primary=1 if i == 0 else 0
                    ))
                result["added"] += 1
            
            row_count += 1
        
        result["total"] = row_count
        
        session.add(ImportHistory(
            filename=os.path.basename(filepath),
            records_added=result["added"],
            records_updated=result["updated"],
            records_skipped=result["skipped"],
        ))
        session.commit()
        
    except Exception as e:
        session.rollback()
        raise RuntimeError(f"Import failed for {filepath}: {e}")
    finally:
        session.close()
    
    return result


# ============================================================================
# AUTO-DETECT FILE FORMAT
# ============================================================================

def import_excel(filepath: str) -> Dict[str, int]:
    """
    Auto-detect the Excel file format and import accordingly.
    
    Detection logic:
      - Check for "NAME" column header → Newly Promoted format
      - Check for "Last Name" column header → NUC format
      - Otherwise, show error
    
    Args:
        filepath: Path to the .xlsx file
    
    Returns:
        Dict with import results
    """
    if not os.path.exists(filepath):
        raise FileNotFoundError(f"File not found: {filepath}")
    
    if not filepath.lower().endswith('.xlsx'):
        raise ValueError("Only .xlsx files are supported. Please use the NUC template format.")
    
    # Detect format by peeking at the first few rows
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active
    
    # Scan first 8 rows for identifying column names
    header_text = ""
    for i in range(1, min(9, ws.max_row + 1)):
        for col in range(1, ws.max_column + 1):
            val = str(ws.cell(row=i, column=col).value or "").strip()
            header_text += val.upper() + " "
    
    if "LAST NAME" in header_text:
        return import_nuc_format(filepath)
    elif "NAME" in header_text and "SEX" in header_text:
        return import_newly_promoted(filepath)
    else:
        raise ValueError(
            "Unrecognized Excel format. Expected either:\n"
            "  1. NUC format: columns with 'Last Name', 'First Name', etc.\n"
            "  2. Newly Promoted format: columns with 'NAME', 'SEX', etc.\n\n"
            f"Detected headers: {header_text[:200]}..."
        )


# ============================================================================
# EXPORT TO NUC FORMAT
# ============================================================================

def export_nuc_format(
    output_path: str,
    include_retired: bool = False,
    sort_by: str = "faculty"
) -> int:
    """
    Export the database to NUC format Excel file.
    
    Generates sequential S/No. based on current sort order.
    Multiple phone numbers are joined with ' / ' delimiter.
    Footer includes attribution to the Directorate of Academic Planning.
    
    Args:
        output_path: Where to save the .xlsx file
        include_retired: If True, include retired professors
        sort_by: Sort order — "faculty" (default), "department", "name"
    
    Returns:
        Number of professors exported
    """
    session = get_session()
    
    try:
        # Query professors
        query = session.query(Professor)
        if not include_retired:
            query = query.filter_by(is_retired=0)
        
        # Sort
        if sort_by == "name":
            query = query.order_by(Professor.last_name, Professor.first_name)
        elif sort_by == "department":
            query = query.order_by(Professor.department, Professor.last_name)
        else:  # default: by faculty
            query = query.order_by(Professor.faculty, Professor.department, Professor.last_name)
        
        professors = query.all()
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Full Professors List"
        
        # Styles
        header_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
        header_fill = PatternFill(start_color="00843D", end_color="00843D", fill_type="solid")  # ATBU Green
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell_alignment = Alignment(vertical="top", wrap_text=True)
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        
        # Row 1: Title
        ws.merge_cells("A1:J1")
        title_cell = ws["A1"]
        title_cell.value = "ABUBAKAR TAFAWA BALEWA UNIVERSITY, BAUCHI — FULL PROFESSORS LIST"
        title_cell.font = Font(name="Calibri", bold=True, size=14, color="00843D")
        title_cell.alignment = Alignment(horizontal="center")
        ws.row_dimensions[1].height = 30
        
        # Row 2: Column headers (matching NUC format)
        headers = [
            "S/No.", "Last Name", "First Name", "Other Names",
            "Date of Professorship", "Faculty", "Department",
            "Area of Specialization", "E-mail", "Phone Number"
        ]
        
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        ws.row_dimensions[2].height = 25
        
        # Data rows
        for i, prof in enumerate(professors, 1):
            row_num = i + 2
            
            # Phone numbers: join multiple with ' / '
            phone_text = " / ".join(pn.phone for pn in prof.phone_numbers)
            
            values = [
                i,  # S/No. (regenerated)
                prof.last_name,
                prof.first_name,
                prof.other_names or "",
                prof.date_of_professorship,
                prof.faculty,
                prof.department,
                prof.area_of_specialization,
                prof.email or "",
                phone_text,
            ]
            
            for col_idx, value in enumerate(values, 1):
                cell = ws.cell(row=row_num, column=col_idx, value=value)
                cell.alignment = cell_alignment
                cell.border = thin_border
                cell.font = Font(name="Calibri", size=10)
            
            # Highlight retired professors (strikethrough, grey)
            if prof.is_retired:
                for col_idx in range(1, 11):
                    cell = ws.cell(row=row_num, column=col_idx)
                    cell.font = Font(name="Calibri", size=10, color="999999")
        
        # Column widths
        col_widths = [8, 22, 20, 22, 22, 30, 30, 35, 30, 25]
        for i, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
        
        # Footer row with attribution
        footer_row = len(professors) + 4
        ws.merge_cells(f"A{footer_row}:J{footer_row}")
        footer_cell = ws.cell(row=footer_row, column=1)
        footer_cell.value = (
            "Generated by the ATBU Academic Planning Portal — "
            "Directorate of Academic Planning | "
            f"Exported: {datetime.now().strftime('%Y-%m-%d %H:%M')} | "
            f"Includes {len(professors)} professors"
        )
        footer_cell.font = Font(name="Calibri", italic=True, size=9, color="666666")
        footer_cell.alignment = Alignment(horizontal="center")
        
        # Save
        wb.save(output_path)
        return len(professors)
        
    finally:
        session.close()
